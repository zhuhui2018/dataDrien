#encoding:utf-8

import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font,colors

class ParseExcel(object):
    def __init__(self,filePath):
        self.filePath = filePath
        self.workbook = load_workbook(filePath)  #加载excel
        self.sheet = self.workbook.active        #获取第一个sheet
        self.font = Font(color=None)
        self.colorDict = {"red": 'FFFF3030', "green": 'FF008B00'}

    #设置当前要操作的sheet对象，使用index来获取相应的sheet
    def get_sheet_by_index(self,sheet_index):
        sheet_name = self.workbook.sheetnames[sheet_index]
        self.sheet = self.get_sheet_by_name(sheet_name)
        return self.sheet

    #设置当前要操作的sheet对象，通过sheet名字来获取sheet
    def get_sheet_by_name(self,sheetName):
        self.sheet = self.workbook[sheetName]
        return self.sheet

    #获取当前默认sheet的名字
    def get_default_sheet(self):
        return self.sheet.title

    #获取默认sheet中最大的行数
    def get_max_row_no(self):
        return self.sheet.max_row

    #获取默认sheet中最大的列数
    def get_max_col_no(self):
        return self.sheet.max_column

    #获取默认sheet中最小的行数
    def get_min_row_no(self):
        return self.sheet.min_row

    #获取默认sheet中最小的列数
    def get_min_row_no(self):
        return self.sheet.min_column

    #获取默认sheet中的所有行对象
    def get_all_rows(self):
        return list(self.sheet.rows)

    #获取默认sheet中的所有列对象
    def get_all_columns(self):
        return list(self.sheet.columns)

    #获取默认sheet中某一行，第一行从0开始
    def get_row(self,row_no):
        return list(self.sheet.rows)[row_no-1]

    #获取最大行
    def get_max_row(self):
        return self.sheet.max_row

    # 获取默认sheet中某一列，第一列从0开始
    def get_col(self,col_no):
        return list(self.sheet.columns)[col_no-1]

    #默认sheet中，通过行号和列号获取指定单元格的内容，行号和列号从1开始
    def get_cell(self,row_no,col_no):
        return self.sheet.cell(row=row_no,column=col_no)

    #默认sheet中，通过行号和列号获取指定单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self,row_no,col_no):
        return self.sheet.cell(row=row_no,column=col_no).value

    def get_cell_content_by_coordinate(self,coordinate):
        return self.sheet[coordinate].value

    #从默认sheet中，通过行号和列号向指定单元格中写入指定内容，行号和列号从1开始
    def write_cell_content(self,row_no,col_no,content,font=None,color=None):
        try:
            self.sheet.cell(row=row_no,column=col_no).value=content
            if color:
                #color填指定的颜色比如red，green即可
                self.sheet.cell(row=row_no,column=col_no).value=content
                self.sheet.cell(row=row_no,column=col_no).value=content
                self.sheet.cell(row=row_no,column=col_no).font=Font(color=self.colorDict[color])
                self.workbook.save(self.filePath)
            else:
                self.sheet.cell(row=row_no, column=col_no).font = self.font
            self.workbook.save(self.filePath)
            return self.sheet.cell(row=row_no,column=col_no).value
        except Exception as e:
            raise e

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期
    def write_cell_current_time(self,row_no,col_no):
        current_time = datetime.datetime.now()
        self.sheet.cell(row=row_no,column=col_no).value=str(current_time)
        self.workbook.save(self.filePath)
        return self.sheet.cell(row=row_no,column=col_no).value

    def save_excel_file(self):
        self.workbook.save(self.filePath)

if __name__=="__main__":
    pe=ParseExcel(r"D:\126邮箱联系人.xlsx")
    pe.get_sheet_by_name("联系人")
    print(pe.get_max_row_no())
    print(pe.write_cell_content(5, 7, "测试自家",color="red"))
    print(pe.write_cell_current_time(5,8))











